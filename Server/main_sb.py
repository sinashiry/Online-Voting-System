# -*- coding: utf-8 -*-

#####################################################
#    Project:     Online Voting System v2           #
#    Designer:    Hosein AlamShahi                  #
#    Programmer:  Sina Shiry                        #
#    Date:        2019 Mar 03                       #
#    For:         Refinery Of Tabriz                #
#####################################################


# Imported Modules
from PyQt5 import QtCore, QtGui, QtWidgets
import sys

import ui_basic
import ui_add_new
import ui_add_old
import ui_active_old
import ui_about
import ui_ip
import ui_server_ip
import ui_report

import sqlite3
import pyodbc
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

import os
from os.path import exists

import time
import threading
import sys

from matplotlib.backends.backend_qt5agg import FigureCanvas
#from matplotlib.figure import Figure
import matplotlib.pyplot as plt

from bidi import algorithm as bidialg
import arabic_reshaper


#### Global Variables ####
poll_IP = "Poll_IP.db"
server_IP = "Server_IP.db"
chart_clc = True
firstTime = False
poll_now = "0"

#### GLOBAL INFO> ABOUT SERVER ####
fs_ip = open("server.txt", 'r')
sqlto_line = fs_ip.readline()
updateTime_line = fs_ip.readline()

sqlTimeout = sqlto_line[3:]
sqlTimeout = sqlTimeout.replace("\n","")
sqlTimeout = sqlTimeout.replace(" ","")

sqlTimeout = int(sqlTimeout)
updateTime = updateTime_line.replace("\n","")
updateTimeinSec = int(updateTime[0:updateTime.find(":")]) * 60 + int(updateTime[updateTime.find(":")+1:])


choice_ = [bidialg.get_display(arabic_reshaper.reshape(u'گزینه یک')),
        bidialg.get_display(arabic_reshaper.reshape(u'گزینه دو')),
        bidialg.get_display(arabic_reshaper.reshape(u'گزینه سه')),
        bidialg.get_display(arabic_reshaper.reshape(u'گزینه چهار')),
        bidialg.get_display(arabic_reshaper.reshape(u'گزینه پنج'))]


buttons = {'BUTTON_1':'گزینه یک', 'BUTTON_2':'گزینه دو',
            'BUTTON_3':'گزینه سه', 'BUTTON_4':'گزینه چهار', 'BUTTON_5':'گزینه پنج'}

deviceNameBase = {'Device_1':'دستگاه یک', 'Device_2':'دستگاه دو',
            'Device_3':'دستگاه سه', 'Device_4':'دستگاه چهار',
            'Device_5':'دستگاه پنج'}

deviceNameBaseReverse = {'دستگاه یک':'Device_1','دستگاه دو':'Device_2',
            'دستگاه سه':'Device_3','دستگاه چهار':'Device_4',
            'دستگاه پنج':'Device_5'}

# Report section list(s) and variables
smallestYear = 1397
biggestYear = 1500


### Section about MAIN Window


#/\ Classes

class thread_ping(QtCore.QThread):
    sig = QtCore.pyqtSignal(int)
    sig2 = QtCore.pyqtSignal(int)
    def __init__(self):
        super(thread_ping,self).__init__()
        self.sig.connect(set_onoff_mode)
        self.sig2.connect(update_data) 
    def run(self):
        i = 0
        global updateTimeJ
        updateTimeJ = 0
        while(True):
            if (i == 15):
                self.sig.emit(set_onoff_mode)
                i = 0
            if (updateTimeJ == updateTimeinSec):
                if (ui_main.autoUpdateCheckBox.isChecked()):
                    self.sig2.emit(update_data)
                updateTimeJ = 0
            self.sleep(1)
            i = i + 1
            updateTimeJ = updateTimeJ + 1


# Functions
#/\
def create_main_folder():
    if not(exists(poll_IP)):
        conn_poll_db = sqlite3.connect(poll_IP)
        conn_poll_db.execute('''CREATE TABLE POLL_ADDRESSES(
                IP_1    TEXT    NOT NULL,
                IP_2    TEXT    NOT NULL,
                IP_3    TEXT    NOT NULL,
                IP_4    TEXT    NOT NULL,
                IP_5    TEXT    NOT NULL);''')
        conn_poll_db.close()

        IP_FT = ['0.0.0.0','0.0.0.0','0.0.0.0','0.0.0.0','0.0.0.0']

        conn_poll_db = sqlite3.connect(poll_IP)
        conn_poll_db.execute("INSERT INTO POLL_ADDRESSES (IP_1,IP_2,IP_3,IP_4,IP_5) " +
            " VALUES ('"+IP_FT[0]+"','"+IP_FT[1]+"','"+IP_FT[2]+"','"+IP_FT[3]+"','"+IP_FT[4]+"')")
        conn_poll_db.commit()
        conn_poll_db.close()
    if not(exists(server_IP)):
        conn_poll_db = sqlite3.connect(server_IP)
        conn_poll_db.execute('''CREATE TABLE SERVER_ADDRESS(
                IP    TEXT    NOT NULL,
                PO    TEXT    NOT NULL,
                US    TEXT    NOT NULL,
                PA    TEXT    NOT NULL,
                DB    TEXT    NOT NULL);''')
        conn_poll_db.close()

        IP_FT = ['0.0.0.0','1433','sa','','voting']

        conn_poll_db = sqlite3.connect(server_IP)
        conn_poll_db.execute("INSERT INTO SERVER_ADDRESS (IP,PO,US,PA,DB) " +
            " VALUES ('"+IP_FT[0]+"','"+IP_FT[1]+"','"+IP_FT[2]+"','"+IP_FT[3]+"','"+IP_FT[4]+"')")
        conn_poll_db.commit()
        conn_poll_db.close()


#/\ Get Client Addresses from Local DB
def get_ip_from_db():
    #Check to DB file exists
    create_main_folder()

    addrss = sqlite3.connect(poll_IP)
    cb = addrss.cursor()
    cb.execute('SELECT * FROM {tn}'.\
                    format(tn="POLL_ADDRESSES"))
    addrss_ = cb.fetchall()
    addrss.close()

    global client_address
    client_address = [str(addrss_[0][0]),str(addrss_[0][1]),str(addrss_[0][2]),str(addrss_[0][3]),str(addrss_[0][4])]

    global device_
    device_ = {client_address[0]:'دستگاه یک', client_address[1]:'دستگاه دو',
            client_address[2]:'دستگاه سه', client_address[3]:'دستگاه چهار',
            client_address[4]:'دستگاه پنج'}
    global deviceReverse
    deviceReverse = {'دستگاه یک':client_address[0],'دستگاه دو':client_address[1],
            'دستگاه سه':client_address[2],'دستگاه چهار':client_address[3],
            'دستگاه پنج':client_address[4]}


#/\
def find_lastActive_table():
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable != None:
            cur.execute("SELECT PATH_NUM FROM POLL_QUESTION WHERE ACT_VOTE = '*'")
            combo_ = cur.fetchall()[0][0]
            global poll_now
            poll_now = str(combo_)
            poll_now = int(poll_now)
            poll_now = str(poll_now)
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText(" نظرسنجی ایجاد نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()


#/\
def find_ips(combo_mode):
    if combo_mode == "all":
        ips = list(deviceNameBase)
        return ips
    else:
        for ip,device_name in deviceNameBase.items():
            if ui_main.comboBox.currentText() == device_name:
                break
        global ip_now
        ip_now = ip
        ips = [""]
        ips[0] = ip
        return ips

#/\
def update_labels():
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable != None:
            cur.execute("SELECT * FROM POLL_QUESTION WHERE PATH_NUM = '" + poll_now + "'")
            labels = cur.fetchall()
            #update labels text
            ui_main.question.setText(labels[0][0])
            ui_main.label_2.setText(labels[0][1])
            ui_main.label_3.setText(labels[0][2])
            ui_main.label_4.setText(labels[0][3])
            ui_main.label_5.setText(labels[0][4])
            ui_main.label_6.setText(labels[0][5])
            #update Button Dict
            global buttons
            buttons = {'BUTTON_1':labels[0][1], 'BUTTON_2':labels[0][2],
                    'BUTTON_3':labels[0][3], 'BUTTON_4':labels[0][4], 'BUTTON_5':labels[0][5]}
            
            global choice_
            choice_ = [bidialg.get_display(arabic_reshaper.reshape(buttons['BUTTON_1'])),
                    bidialg.get_display(arabic_reshaper.reshape(buttons['BUTTON_2'])),
                    bidialg.get_display(arabic_reshaper.reshape(buttons['BUTTON_3'])),
                    bidialg.get_display(arabic_reshaper.reshape(buttons['BUTTON_4'])),
                    bidialg.get_display(arabic_reshaper.reshape(buttons['BUTTON_5']))]
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText(" نظرسنجی ایجاد نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()


#/\
def update_statistics(path, combo_mode):
    all_rows = list()
    BUTTON_COUNT = [0, 0, 0, 0, 0]
    for i in path:
        #open Database
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        cur.execute("SELECT VOTE,DATE,TIME,DIVC FROM vote_" + poll_now + " WHERE DIVC = '" + i + "' ORDER BY DATE, TIME")
        all_rows = all_rows + cur.fetchall()
    ### Get Data from DataBase About Choices
    if combo_mode == "all":
        for but_num in range(1,6):
            cur.execute("SELECT COUNT(VOTE) FROM vote_" + poll_now + " WHERE VOTE = 'BUTTON_" + str(but_num) + "'")
            cache_ = cur.fetchall()
            BUTTON_COUNT[but_num-1] = cache_[0][0]
    else:
        for but_num in range(1,6):
            cur.execute("SELECT COUNT(VOTE) FROM vote_" + poll_now + " WHERE VOTE = 'BUTTON_" + str(but_num) + "' AND DIVC = '" + path[0] + "'")
            cache_ = cur.fetchall()
            BUTTON_COUNT[but_num-1] = cache_[0][0]
    if len(all_rows) == 0:
        BUTTON_COUNT = [0, 0, 0, 0, 0]
        if (combo_mode == "all"):
            warn = QtWidgets.QMessageBox()
            warn.setText("نظری توسط هیچیک از دستگاه ها ثبت نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText("نظری توسط این دستگاه ثبت نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    
    ### Create QT Table for show Vote(s)
    if combo_mode == "all":
        col_ = 4
    else:
        col_ = 3 
    ui_main.tableWidget.setColumnCount(col_)
    ui_main.tableWidget.setRowCount(len(all_rows))
    
    ### Set Header Of table
    if combo_mode == "all":
        ui_main.tableWidget.setHorizontalHeaderLabels(['گزینه انتخابی','تاریخ','ساعت','دستگاه'])
    else:
        ui_main.tableWidget.setHorizontalHeaderLabels(['گزینه انتخابی','تاریخ','ساعت'])

    #Not Editable Table
    ui_main.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
    
    ### Put data in QT Table
    for x in range(0,len(all_rows)):
        for y in range(0,col_):
            now_txt = all_rows[x][y]
            
            if y == 0:
                now_txt = buttons.get(now_txt)
            elif y == 3:
                now_txt = deviceNameBase.get(now_txt)
            else:
                None
            
            ui_main.tableWidget.setItem(x,y, QtWidgets.QTableWidgetItem(now_txt))
    
    ### Update each of Choices
    ui_main.LCD_NUM_1.display(BUTTON_COUNT[0])
    ui_main.LCD_NUM_2.display(BUTTON_COUNT[1])
    ui_main.LCD_NUM_3.display(BUTTON_COUNT[2])
    ui_main.LCD_NUM_4.display(BUTTON_COUNT[3])
    ui_main.LCD_NUM_5.display(BUTTON_COUNT[4])
    ### Update Row length of DataBase
    ui_main.LCD_NUM_TOTAL.display(len(all_rows))

    ### Update Pie Chart
    if chart_clc == False:
        dele()
    # Only Pies that greater than 
    BUTTON_COUNT_Chart = list()
    CHOICE_CHART = list()
    j = 0
    for i in range(1,6):
        if BUTTON_COUNT[i-1] != 0 :
            BUTTON_COUNT_Chart.append(BUTTON_COUNT[i-1])
            CHOICE_CHART.append(choice_[i-1])
            j += 1
        else:
            continue

    fig = plt.figure(1, figsize=(10,10))
    fig.patch.set_facecolor('None')
    fig.patch.set_alpha(0)
    fig.add_subplot(111)
    #color = 'b', 'o', 'g', 'r', 'p'
    plt.pie(BUTTON_COUNT_Chart, labels=CHOICE_CHART, autopct='%1.1f%%', shadow=True, startangle=90)
    global static_canvas
    static_canvas = FigureCanvas(fig)
    layout.addWidget(static_canvas)

    # Show the Last vote Pic
    if (len(all_rows) != 0 and combo_mode == "single"):
        cur.execute("SELECT IMAG FROM vote_" + poll_now + " WHERE DIVC = '" + path[0] + "' ORDER BY DATE, TIME")
        r = cur.fetchall()
        r = r[len(r)-1][0]
        dataCache = r
        ff = open("cache.jpg", mode='wb')
        ff.write(dataCache)
        ff.close()
        pic = QtGui.QPixmap("cache.jpg")
        scene = QtWidgets.QGraphicsScene()
        scene.addPixmap(pic)
        ui_main.graphicsView.fitInView(scene.itemsBoundingRect())
        ui_main.graphicsView.setScene(scene)
        ui_main.graphicsView.show()
    elif (len(all_rows) != 0 and combo_mode == "all"):
        cur.execute("SELECT MAX(DATE) FROM vote_" + poll_now )
        rDate = cur.fetchall()
        rDate = rDate[0][0]
        cur.execute("SELECT MAX(TIME) FROM vote_" + poll_now + " WHERE DATE = '" + rDate + "'")
        rTime = cur.fetchall()
        rTime = rTime[0][0]
        cur.execute("SELECT IMAG FROM vote_" + poll_now + " WHERE TIME = '" + rTime + "' AND DATE = '" + rDate + "' ORDER BY DATE, TIME")
        r = cur.fetchall()
        r = r[len(r)-1][0]
        dataCache = r
        ff = open("cache.jpg", mode='wb')
        ff.write(dataCache)
        ff.close()
        pic = QtGui.QPixmap("cache.jpg")
        scene = QtWidgets.QGraphicsScene()
        scene.addPixmap(pic)
        ui_main.graphicsView.fitInView(scene.itemsBoundingRect())
        ui_main.graphicsView.setScene(scene)
        ui_main.graphicsView.show()
    else:
        pic = QtGui.QPixmap("nophoto.jpg")
        scene = QtWidgets.QGraphicsScene()
        scene.addPixmap(pic)
        ui_main.graphicsView.fitInView(scene.itemsBoundingRect())
        ui_main.graphicsView.setScene(scene)
        ui_main.graphicsView.show()
    #close DataBase File
    cur.close()

#/\
def dele():
    plt.close()
    static_canvas.setParent(None)


#/\
def update_data():
    global chart_clc
    # start progress bar
    if (ui_main.comboBox.currentText() == "جمع کل"):
        #try to connect sql server
        try:
            connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
            cur = connnection.cursor()
            # check exists of Table
            tableName = "vote_" + poll_now
            checkTable = cur.tables(table=tableName).fetchone()
            cur.close()
            if checkTable == None:
                warn = QtWidgets.QMessageBox()
                warn.setText("جدول داده مربوطه در سرور یافت نشد.")
                warn.setWindowTitle("اخطار")
                warn.exec_()
            else:
                ips = find_ips("all")
                update_statistics(ips, "all")
                chart_clc = False
        except:
            warn = QtWidgets.QMessageBox()
            warn.setText("امکان اتصال به سرور وجود ندارد")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    else:
        try:
            #try to connect sql server
            connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
            cur = connnection.cursor()
            # check exists of Table
            tableName = "vote_" + poll_now
            checkTable = cur.tables(table=tableName).fetchone()
            cur.close()
            if checkTable == None:
                warn = QtWidgets.QMessageBox()
                warn.setText("جدول داده مربوطه در سرور یافت نشد.")
                warn.setWindowTitle("اخطار")
                warn.exec_()
            else:
                ips = find_ips("single")
                update_statistics(ips, "single")
                chart_clc = False
        except:
            warn = QtWidgets.QMessageBox()
            warn.setText("امکان اتصال به سرور وجود ندارد")
            warn.setWindowTitle("اخطار")
            warn.exec_()

#/\ Global variable that need this Fun.
def show_pic():
    # build pic path and name
    rRow = ui_main.tableWidget.currentRow()
    rIp = ""
    if ui_main.comboBox.currentText() == "جمع کل":
        rIp = ui_main.tableWidget.item(rRow,3).text()
        rIp = deviceNameBaseReverse.get(rIp)
    else:
        rIp = deviceNameBaseReverse.get(ui_main.comboBox.currentText())
    rDate = ui_main.tableWidget.item(rRow,1).text()
    rTime = ui_main.tableWidget.item(rRow,2).text()
    
    #Try to connect SERVER and Get Image
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        cur.execute("SELECT IMAG FROM vote_" + poll_now + " WHERE TIME = '" + rTime + "' AND DATE = '" + rDate + "' AND DIVC = '" + rIp + "'")
        r = cur.fetchall()
        
        if (len(r) == 0):
            warn = QtWidgets.QMessageBox()
            warn.setText("عکس مربوط به نظر انتخاب شده یافت نشد یا دیتا بیس به درستی انتخاب نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
        else:
            r = r[0][0]
            dataCache = r
            ff = open("cache.jpg", mode='wb')
            ff.write(dataCache)
            ff.close()
            pic = QtGui.QPixmap("cache.jpg")
            scene = QtWidgets.QGraphicsScene()
            scene.addPixmap(pic)
            ui_main.graphicsView.fitInView(scene.itemsBoundingRect())
            ui_main.graphicsView.setScene(scene)
            ui_main.graphicsView.show()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()

#/\
def show_pic_win():
    if (not(exists("cache.jpg")) or ui_main.tableWidget.rowCount() == 0):
        warn = QtWidgets.QMessageBox()
        warn.setText("عکس مربوط به نظر انتخاب شده یافت نشد یا دیتا بیس به درستی انتخاب نشده است")
        warn.setWindowTitle("اخطار")
        warn.exec_()
    else:    
        os.startfile("cache.jpg")

#/\
def set_onoff_mode():
    ips = client_address
    rpis_mode = ['off', 'off', 'off', 'off', 'off']
    for i in range(0,5):
        command = "ping -w 1 -n 1 " + ips[i]
        output = os.popen(command).read()
        if ('100% loss' in str(output)):
            rpis_mode[i] = 'off'
        else:
            rpis_mode[i] = 'on'
    
    
    pic_off = QtGui.QPixmap("offline.png")
    pic_on = QtGui.QPixmap("online.png")
    
    scene_off = QtWidgets.QGraphicsScene()
    scene_off.addPixmap(pic_off)
    scene_on = QtWidgets.QGraphicsScene()
    scene_on.addPixmap(pic_on)
    
    if rpis_mode[0]=='on':
        ui_main.rpi_mode_1.setScene(scene_on)
    else:
        ui_main.rpi_mode_1.setScene(scene_off)
    
    if rpis_mode[1]=='on':
        ui_main.rpi_mode_2.setScene(scene_on)
    else:
        ui_main.rpi_mode_2.setScene(scene_off)
    
    if rpis_mode[2]=='on':
        ui_main.rpi_mode_3.setScene(scene_on)
    else:
        ui_main.rpi_mode_3.setScene(scene_off)
    
    if rpis_mode[3]=='on':
        ui_main.rpi_mode_4.setScene(scene_on)
    else:
        ui_main.rpi_mode_4.setScene(scene_off)
    
    if rpis_mode[4]=='on':
        ui_main.rpi_mode_5.setScene(scene_on)
    else:
        ui_main.rpi_mode_5.setScene(scene_off)
    
    ui_main.rpi_mode_1.show()
    ui_main.rpi_mode_2.show()
    ui_main.rpi_mode_3.show()
    ui_main.rpi_mode_4.show()
    ui_main.rpi_mode_5.show()

## Section about Add New Window

# Functions
def create_poll_db_combobox():
    # Get INFO. from Lines
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable != None:
            cur.execute("SELECT * FROM POLL_QUESTION")
            allrows = cur.fetchall()
            for i in range(0,len(allrows)):
                ui_add_o.question_box.addItem(allrows[i][0])
                if (i == 0):
                    ui_add_o.labelButton1.setText(allrows[i][1])
                    ui_add_o.labelButton2.setText(allrows[i][2])
                    ui_add_o.labelButton3.setText(allrows[i][3])
                    ui_add_o.labelButton4.setText(allrows[i][4])
                    ui_add_o.labelButton5.setText(allrows[i][5])
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText(" نظرسنجی ایجاد نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()


# Start add Old Poll_System in main Program and Database
def start_add_old():
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable != None:
            global OWindow
            global ui_add_o
            OWindow = QtWidgets.QMainWindow()
            ui_add_o = ui_add_old.Ui_OWindow()
            ui_add_o.setupUi(OWindow)
            ui_add_o.ok_button.clicked.connect(add_old_poll)
            ui_add_o.cancel_button.clicked.connect(cancel_add_old)
            ui_add_o.question_box.activated.connect(update_old_combo)
            OWindow.show()
            create_poll_db_combobox()
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText(" نظرسنجی ایجاد نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()


def add_old_poll():
    #clear and Hide Window
    cancel_add_old()
    
    # Find Path_number acording to Selected Question

    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable != None:
            cur.execute("SELECT PATH_NUM FROM POLL_QUESTION WHERE QUESTION = '" + ui_add_o.question_box.currentText() + "'")
            combo_ = cur.fetchall()[0][0]
            global poll_now
            poll_now = str(combo_)
            poll_now = int(poll_now)
            poll_now = str(poll_now)
            # Update Labels
            update_labels()
            update_data()
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText(" نظرسنجی ایجاد نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()


def cancel_add_old():
    #Hide Window
    OWindow.hide()

def update_old_combo():
    global OWindow
    global ui_add_o
    connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
    cur = connnection.cursor()
    cur.execute("SELECT * FROM POLL_QUESTION WHERE QUESTION = '" + ui_add_o.question_box.currentText() + "'")
    allrows = cur.fetchall()
    ui_add_o.labelButton1.setText(allrows[0][1])
    ui_add_o.labelButton2.setText(allrows[0][2])
    ui_add_o.labelButton3.setText(allrows[0][3])
    ui_add_o.labelButton4.setText(allrows[0][4])
    ui_add_o.labelButton5.setText(allrows[0][5])

## Section about Add New Window

# Functions
def create_poll_db():
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable == None:
            commnd = "CREATE TABLE dbo.POLL_QUESTION ("
            commnd += "QUESTION    nvarchar(150),"
            commnd += "CHOICE_1    nvarchar(50),"
            commnd += "CHOICE_2    nvarchar(50),"
            commnd += "CHOICE_3    nvarchar(50),"
            commnd += "CHOICE_4    nvarchar(50),"
            commnd += "CHOICE_5    nvarchar(50),"
            commnd += "PATH_NUM    nvarchar(50),"
            commnd += "ACT_VOTE    nvarchar(50))"
            cur.execute(commnd)
            cur.commit()
            cur.close()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()


def add_new_poll():
    # create or open DB
    create_poll_db()
    # Get INFO. from Lines
    QUESTION_val = ui_add_n.Question.text()
    CHOICE1_val = ui_add_n.Choice_1.text()
    CHOICE2_val = ui_add_n.Choice_2.text()
    CHOICE3_val = ui_add_n.Choice_3.text()
    CHOICE4_val = ui_add_n.Choice_4.text()
    CHOICE5_val = ui_add_n.Choice_5.text()
    global poll_now
    PATH_val = str(int(poll_now) + 1)
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        commnd = "CREATE TABLE dbo.vote_" + PATH_val + "("
        commnd += "VOTE    varchar(50),"
        commnd += "DATE    varchar(50),"
        commnd += "TIME    varchar(50),"
        commnd += "DIVC    varchar(50),"
        commnd += "IMAG    image,"
        commnd += "TIDA    DateTimeOffset NOT NULL default SYSDATETIMEOFFSET())"
        cur.execute(commnd)
        cur.commit()
        #REMOVE ACTIVE TABLE
        commnd = "UPDATE POLL_QUESTION SET ACT_VOTE = ''"
        cur.execute(commnd)
        cur.commit()
        # SAVE INFO. in DB File
        commnd = "INSERT INTO POLL_QUESTION (QUESTION,CHOICE_1,CHOICE_2,CHOICE_3,CHOICE_4,CHOICE_5,PATH_NUM,ACT_VOTE)"
        commnd += " VALUES ('"+QUESTION_val+"','"+CHOICE1_val+"','"+CHOICE2_val+"','"+CHOICE3_val+"','"+CHOICE4_val+"','"+CHOICE5_val+"','"+PATH_val+"','" + "*" + "')"
        cur.execute(commnd)
        cur.commit()
        #Update current Voting number
        poll_now = PATH_val
        # Update Labels
        update_labels()
        update_data()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()
    #clear and Hide Window
    cancel_add_new()

def cancel_add_new():
    #clear and Hide Window
    ui_add_n.Question.clear()
    ui_add_n.Choice_1.clear()
    ui_add_n.Choice_2.clear()
    ui_add_n.Choice_3.clear()
    ui_add_n.Choice_4.clear()
    ui_add_n.Choice_5.clear()
    NWindow.hide()


### Setup UI Basic Settings ###
# Start Add New Poll_System in Main Program and Database
def start_add_new():
    global NWindow
    NWindow = QtWidgets.QMainWindow()
    global ui_add_n
    ui_add_n = ui_add_new.Ui_NWindow()
    ui_add_n.setupUi(NWindow)

    ui_add_n.ok_button.clicked.connect(add_new_poll)
    ui_add_n.cancel_button.clicked.connect(cancel_add_new)
    
    ui_add_n.Question.setMaxLength(40)
    ui_add_n.Choice_1.setMaxLength(11)
    ui_add_n.Choice_2.setMaxLength(11)
    ui_add_n.Choice_3.setMaxLength(11)
    ui_add_n.Choice_4.setMaxLength(11)
    ui_add_n.Choice_5.setMaxLength(11)

    NWindow.show()


# Start active Old Poll_System in main Program and Database
def start_active_old():
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable != None:
            global AOWindow
            global ui_active_o
            AOWindow = QtWidgets.QMainWindow()
            ui_active_o = ui_active_old.Ui_AOWindow()
            ui_active_o.setupUi(AOWindow)
            ui_active_o.ok_button.clicked.connect(active_old_poll)
            ui_active_o.cancel_button.clicked.connect(cancel_active_old)
            ui_active_o.question_box.activated.connect(update_active_combo)
            AOWindow.show()
            cur.execute("SELECT * FROM POLL_QUESTION")
            allrows = cur.fetchall()
            for i in range(0,len(allrows)):
                ui_active_o.question_box.addItem(allrows[i][0])
                if (i == 0):
                    ui_active_o.labelButton1.setText(allrows[i][1])
                    ui_active_o.labelButton2.setText(allrows[i][2])
                    ui_active_o.labelButton3.setText(allrows[i][3])
                    ui_active_o.labelButton4.setText(allrows[i][4])
                    ui_active_o.labelButton5.setText(allrows[i][5])
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText(" نظرسنجی ایجاد نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()


def active_old_poll():
    #clear and Hide Window
    cancel_active_old()
    
    # Find Path_number acording to Selected Question

    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable != None:
            #REMOVE ACTIVE TABLE
            commnd = "UPDATE POLL_QUESTION SET ACT_VOTE = ''"
            cur.execute(commnd)
            cur.commit()
            #ACTIVE SELECTED TABLE
            commnd = "UPDATE POLL_QUESTION SET ACT_VOTE = '*' WHERE QUESTION = '" + ui_active_o.question_box.currentText() + "'"
            cur.execute(commnd)
            cur.commit()
            cur.execute("SELECT PATH_NUM FROM POLL_QUESTION WHERE QUESTION = '" + ui_active_o.question_box.currentText() + "'")
            combo_ = cur.fetchall()[0][0]
            global poll_now
            poll_now = str(combo_)
            poll_now = int(poll_now)
            poll_now = str(poll_now)
            # Update Labels
            update_labels()
            update_data()
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText(" نظرسنجی ایجاد نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()


def cancel_active_old():
    #Hide Window
    AOWindow.hide()


def update_active_combo():
    global AOWindow
    global ui_active_o
    connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
    cur = connnection.cursor()
    cur.execute("SELECT * FROM POLL_QUESTION WHERE QUESTION = '" + ui_active_o.question_box.currentText() + "'")
    allrows = cur.fetchall()
    ui_active_o.labelButton1.setText(allrows[0][1])
    ui_active_o.labelButton2.setText(allrows[0][2])
    ui_active_o.labelButton3.setText(allrows[0][3])
    ui_active_o.labelButton4.setText(allrows[0][4])
    ui_active_o.labelButton5.setText(allrows[0][5])

# Start About Us Window
def start_about():
    global AWindow
    AWindow = QtWidgets.QMainWindow()
    global ui_about_n
    ui_about_n = ui_about.Ui_AWindow()
    ui_about_n.setupUi(AWindow)
    AWindow.show()

# Start Config. Clients Ip Addresses
def client_ip_config():
    global IWindow
    IWindow = QtWidgets.QMainWindow()
    global ui_client_n
    ui_client_n = ui_ip.Ui_IWindow()
    ui_client_n.setupUi(IWindow)
    
    ui_client_n.ok_ip_button.clicked.connect(add_client_ip)
    ui_client_n.cancel_ip_button.clicked.connect(cancel_client_ip)

    ui_client_n.ip_device_1.setMaxLength(15)
    ui_client_n.ip_device_2.setMaxLength(15)
    ui_client_n.ip_device_3.setMaxLength(15)
    ui_client_n.ip_device_4.setMaxLength(15)
    ui_client_n.ip_device_5.setMaxLength(15)
    ui_client_n.ip_device_1.setText(client_address[0])
    ui_client_n.ip_device_2.setText(client_address[1])
    ui_client_n.ip_device_3.setText(client_address[2])
    ui_client_n.ip_device_4.setText(client_address[3])
    ui_client_n.ip_device_5.setText(client_address[4])
    
    IWindow.show()

def add_client_ip():
    # Get INFO. from Lines
    CHOICE1_val = ui_client_n.ip_device_1.text()
    CHOICE2_val = ui_client_n.ip_device_2.text()
    CHOICE3_val = ui_client_n.ip_device_3.text()
    CHOICE4_val = ui_client_n.ip_device_4.text()
    CHOICE5_val = ui_client_n.ip_device_5.text()
    
    # SAVE INFO. in DB File
    conn_poll_db = sqlite3.connect(poll_IP)
    conn_poll_db.execute("UPDATE POLL_ADDRESSES SET IP_1 = '{aa}', IP_2 = '{ab}', IP_3 = '{ac}', IP_4 = '{ad}', IP_5 = '{ae}'".\
                    format(aa=CHOICE1_val,ab=CHOICE2_val,ac=CHOICE3_val,ad=CHOICE4_val,ae=CHOICE5_val))
    conn_poll_db.commit()
    conn_poll_db.close()

    global client_address
    client_address = [CHOICE1_val,CHOICE2_val,CHOICE3_val,CHOICE4_val,CHOICE5_val]
    global device_
    device_ = {client_address[0]:'دستگاه یک', client_address[1]:'دستگاه دو',
            client_address[2]:'دستگاه سه', client_address[3]:'دستگاه چهار',
            client_address[4]:'دستگاه پنج'}
    global deviceReverse
    deviceReverse = {'دستگاه یک':client_address[0],'دستگاه دو':client_address[1],
            'دستگاه سه':client_address[2],'دستگاه چهار':client_address[3],
            'دستگاه پنج':client_address[4]}

    
    #clear and Hide Window
    cancel_client_ip()

def cancel_client_ip():
    #clear and Hide Window
    ui_client_n.ip_device_1.clear()
    ui_client_n.ip_device_2.clear()
    ui_client_n.ip_device_3.clear()
    ui_client_n.ip_device_4.clear()
    ui_client_n.ip_device_5.clear()
    IWindow.hide()

# SERVER IP Settings
def check_default_server_ip():
    addrss = sqlite3.connect(server_IP)
    cb = addrss.cursor()
    cb.execute('SELECT * FROM {tn}'.\
                    format(tn="SERVER_ADDRESS"))
    addrss_ = cb.fetchall()
    addrss.close()

    global firstTime
    if (addrss_[0][3] == ""):
        firstTime = True


def get_server_ip():
    addrss = sqlite3.connect(server_IP)
    cb = addrss.cursor()
    cb.execute('SELECT * FROM {tn}'.\
                    format(tn="SERVER_ADDRESS"))
    addrss_ = cb.fetchall()
    addrss.close()
    
    global server_address
    server_address = [addrss_[0][0],addrss_[0][1],addrss_[0][2],addrss_[0][3],addrss_[0][4]]

    global connSQL
    connSQL = 'DRIVER={SQL Server};SERVER='+server_address[0]+';PORT='+server_address[1]+';DATABASE='+ server_address[4]+';UID='+server_address[2]+';PWD='+server_address[3]


def start_server_ip():
    global SIWindow
    SIWindow = QtWidgets.QMainWindow()
    global ui_server_n
    ui_server_n = ui_server_ip.Ui_SIWindow()
    ui_server_n.setupUi(SIWindow)
    
    ui_server_n.ok_ip_button.clicked.connect(add_server_ip)
    ui_server_n.cancel_ip_button.clicked.connect(cancel_server_ip)

    #Get INFO. from local DB
    addrss = sqlite3.connect(server_IP)
    cb = addrss.cursor()
    cb.execute('SELECT * FROM {tn}'.\
                    format(tn="SERVER_ADDRESS"))
    addrss_ = cb.fetchall()
    addrss.close()
    #Put Info in Inputs rectangles
    #ui_server_n.server_ip.setMaxLength(15)
    ui_server_n.server_ip.setText(addrss_[0][0])
    ui_server_n.server_port.setText(addrss_[0][1])
    ui_server_n.server_user.setText(addrss_[0][2])
    ui_server_n.server_pass.setText(addrss_[0][3])
    ui_server_n.server_db.setText(addrss_[0][4])
    ui_server_n.server_pass.setEchoMode(QtWidgets.QLineEdit.Password)
    ui_server_n.server_db.setDisabled(True)

    SIWindow.show()


def add_server_ip():
    # Get INFO. from Lines
    ipVal = ui_server_n.server_ip.text()
    poVal = ui_server_n.server_port.text()
    usVal = ui_server_n.server_user.text()
    paVal = ui_server_n.server_pass.text()
    dbVal = ui_server_n.server_db.text()

    # SAVE INFO. in DB File
    conn_poll_db = sqlite3.connect(server_IP)
    conn_poll_db.execute("UPDATE SERVER_ADDRESS SET IP = '{aa}', PO = '{ab}', US = '{ac}', PA = '{ad}', DB = '{ae}'".\
                    format(aa=ipVal,ab=poVal,ac=usVal,ad=paVal,ae=dbVal))
    conn_poll_db.commit()
    conn_poll_db.close()

    global server_address
    server_address = [ipVal,poVal,usVal,paVal,dbVal]

    #clear and Hide Window
    cancel_server_ip()

def cancel_server_ip():
    #clear and Hide Window
    ui_server_n.server_ip.clear()
    ui_server_n.server_port.clear()
    ui_server_n.server_user.clear()
    ui_server_n.server_pass.clear()
    ui_server_n.server_db.clear()
    SIWindow.hide()


### AUTO UPDATE 
def autoUpdate():
    #Get Update Time changing
    sec = ui_main.autoUpdatTime.time().second()
    min = ui_main.autoUpdatTime.time().minute()
    updateTime = str(min) + ":" + str(sec)

    global updateTimeinSec
    updateTimeinSec = (min*60 + sec)
    
    resetAutoUpdateJ()

    sqlAndTime = "to=" + str(sqlTimeout) + "\n" + updateTime
    # Save Update Time settings
    file_ = open("server.txt",'w')
    file_.write(sqlAndTime)
    file_.close()


def resetAutoUpdateJ():
    global updateTimeJ
    if ui_main.autoUpdateCheckBox.isChecked():
        updateTimeJ = 0


### REPORT SECTION

#Start report window
def start_report():
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable != None:
            global RWindow
            RWindow = QtWidgets.QMainWindow()
            global ui_report_n
            ui_report_n = ui_report.Ui_RWindow()
            ui_report_n.setupUi(RWindow)
            
            ui_report_n.ok_button.clicked.connect(get_report)
            ui_report_n.cancel_button.clicked.connect(cancel_report)

            #fill year combo boxes
            for i in range(smallestYear,biggestYear):
                ui_report_n.startYear.addItem(str(i))
                ui_report_n.endYear.addItem(str(i))

            #fill month combo boxes
            for i in range(1,13):
                ui_report_n.startMonth.addItem(str(i))
                ui_report_n.endMonth.addItem(str(i))
            
            #fill day combo boxes
            for i in range(1,32):
                ui_report_n.startDay.addItem(str(i))
                ui_report_n.endDay.addItem(str(i))
            
            # Modify days for different months
            ui_report_n.startMonth.activated.connect(changeReportStartDayCombo)
            ui_report_n.endMonth.activated.connect(changeReportEndDayCombo)

            # Save Path
            ui_report_n.savePathLine.setDisabled(True)
            ui_report_n.savePathButton.clicked.connect(getSavePath)

            RWindow.show()
            
            cur.execute("SELECT QUESTION FROM POLL_QUESTION")
            allrows = cur.fetchall()
            for i in range(0,len(allrows)):
                ui_report_n.question_box.addItem(allrows[i][0])
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText(" نظرسنجی ایجاد نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()


def changeReportStartDayCombo():
    if int(ui_report_n.startMonth.currentText()) <= 6:
        ui_report_n.startDay.clear()
        for i in range(1,32):
            ui_report_n.startDay.addItem(str(i))
    else:
        ui_report_n.startDay.clear()
        for i in range(1,31):
            ui_report_n.startDay.addItem(str(i))
def changeReportEndDayCombo():
    if int(ui_report_n.endMonth.currentText()) <= 6:
        ui_report_n.endDay.clear()
        for i in range(1,32):
            ui_report_n.endDay.addItem(str(i))
    else:
        ui_report_n.endDay.clear()
        for i in range(1,31):
            ui_report_n.endDay.addItem(str(i))


def getSavePath():
    #Get user selected path
    savePathDir = QtWidgets.QFileDialog.getExistingDirectoryUrl()
    i = 1
    while(True):
        savePath = savePathDir.path()[1:] + "/output_" + str(i) + ".xlsx"
        if not exists(savePath):
            break
        else:
            i += 1
    #Show Path in line
    ui_report_n.savePathLine.setText(savePath)


# get report file
def get_report():
    try:
        connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
        cur = connnection.cursor()
        checkTable = cur.tables(table="POLL_QUESTION").fetchone()
        if checkTable != None:
            # Get num of voting
            cur.execute("SELECT PATH_NUM FROM POLL_QUESTION WHERE QUESTION = '" + ui_report_n.question_box.currentText() + "'")
            current_path_num = cur.fetchall()[0][0]

            # Build start and end date
            startDate = ui_report_n.startYear.currentText() + "-"
            if int(ui_report_n.startMonth.currentText()) < 10:
                startDate += "0" + ui_report_n.startMonth.currentText()
            else:
                startDate += ui_report_n.startMonth.currentText()
            startDate += "-"
            if int(ui_report_n.startDay.currentText()) < 10:
                startDate += "0" + ui_report_n.startDay.currentText()
            else:
                startDate += ui_report_n.startDay.currentText()

            endDate = ui_report_n.endYear.currentText() + "-"
            if int(ui_report_n.endMonth.currentText()) < 10:
                endDate += "0" + ui_report_n.endMonth.currentText()
            else:
                endDate += ui_report_n.endMonth.currentText()
            endDate += "-"
            if int(ui_report_n.endDay.currentText()) < 10:
                endDate += "0" + ui_report_n.endDay.currentText()
            else:
                endDate += ui_report_n.endDay.currentText()
            
            #Check Date greater than... Condition
            if int(startDate.replace("-", "")) <= int(endDate.replace("-", "")):
                if ui_report_n.savePathLine.text() != "":
                    saveXLSX(startDate,endDate,current_path_num,ui_report_n.savePathLine.text())
                    #clear and Hide Window
                    #cancel_report()
                else:
                    warn = QtWidgets.QMessageBox()
                    warn.setText("مسیر ذخیره انتخاب نشده بود")
                    warn.setWindowTitle("اخطار")
                    warn.exec_()
                    #clear and Hide Window
                    #cancel_report()
            else:
                warn = QtWidgets.QMessageBox()
                warn.setText("تاریخ شروع از تاریخ پایان جلوتر بود")
                warn.setWindowTitle("اخطار")
                warn.exec_()
                #clear and Hide Window
                #cancel_report()
        else:
            warn = QtWidgets.QMessageBox()
            warn.setText(" نظرسنجی ایجاد نشده است")
            warn.setWindowTitle("اخطار")
            warn.exec_()
            #clear and Hide Window
            cancel_report()
    except:
        warn = QtWidgets.QMessageBox()
        warn.setText("امکان اتصال به سرور وجود ندارد")
        warn.setWindowTitle("اخطار")
        warn.exec_()
        #clear and Hide Window
        cancel_report()

#SAVE DATA IN XLSX FILE
def saveXLSX(startDate, endDate, currentVoting, savePathURL):
    connnection = pyodbc.connect(connSQL, timeout = sqlTimeout)
    cur = connnection.cursor()
    cur.execute("SELECT VOTE,DATE,TIME,DIVC FROM vote_" + currentVoting + " WHERE DATE BETWEEN '" + startDate + "' AND '" + endDate + "' ORDER BY DATE, TIME")
    all_rows = cur.fetchall()

    # Get each Voting Buttons
    cur.execute("SELECT CHOICE_1, CHOICE_2, CHOICE_3, CHOICE_4, CHOICE_5 FROM POLL_QUESTION WHERE PATH_NUM = '" + currentVoting + "'")
    all_rows_choices = cur.fetchall()
    
    eachVotingButtons = {"BUTTON_1":all_rows_choices[0][0],
                        "BUTTON_2":all_rows_choices[0][1],
                        "BUTTON_3":all_rows_choices[0][2],
                        "BUTTON_4":all_rows_choices[0][3],
                        "BUTTON_5":all_rows_choices[0][4]}
    
    #Create XLSX File
    if len(all_rows) != 0:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "گزینه انتخابی"; ws['B1'] = "تاریخ"; ws['C1'] = "ساعت"; ws['D1'] = "شماره دستگاه"
        ws['A1'].font = Font(bold=True, italic=True)
        ws['B1'].font = Font(bold=True, italic=True)
        ws['C1'].font = Font(bold=True, italic=True)
        ws['D1'].font = Font(bold=True, italic=True)
        #wb.save(savePathURL)
        
        #Write Data into XLSX File
        for i in range(0, len(all_rows)):
                pos_A = "A" + str(i+2)
                pos_B = "B" + str(i+2)
                pos_C = "C" + str(i+2)
                pos_D = "D" + str(i+2)
                ws[pos_A] = eachVotingButtons[all_rows[i][0]]
                ws[pos_B] = all_rows[i][1]
                ws[pos_C] = all_rows[i][2]
                ws[pos_D] = deviceNameBase[all_rows[i][3]]
        wb.save(savePathURL)

        #Show saved message
        warn = QtWidgets.QMessageBox()
        warn.setText("داده های خواسته شده در مسیر زیر ذخیره شد\n" + savePathURL)
        warn.setWindowTitle("پیغام")
        warn.exec_()
        cancel_report()
    else:
        warn = QtWidgets.QMessageBox()
        warn.setText("در بازه انتخاب شده نظری ثبت نشده است")
        warn.setWindowTitle("اخطار")
        warn.exec_()
        #clear and Hide Window
        #cancel_report()


# canle report window
def cancel_report():
    ui_report_n.startYear.clear()
    ui_report_n.startMonth.clear()
    ui_report_n.startDay.clear()
    ui_report_n.endYear.clear()
    ui_report_n.endMonth.clear()
    ui_report_n.endDay.clear()
    ui_report_n.question_box.clear()
    ui_report_n.savePathLine.clear()
    RWindow.hide()



###### Prebuilding of UI
def prebuild():
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    global ui_main
    ui_main = ui_basic.Ui_MainWindow()
    ui_main.setupUi(MainWindow)

    ui_main.refresh_button.clicked.connect(update_data)
    ui_main.zoom_button.clicked.connect(show_pic_win)
    ui_main.comboBox.activated.connect(update_data)
    ui_main.tableWidget.clicked.connect(show_pic)
    ui_main.new_vote.clicked.connect(start_add_new)
    ui_main.old_vote.clicked.connect(start_add_old)
    ui_main.active_old_vote.clicked.connect(start_active_old)
    ui_main.about_us.clicked.connect(start_about)
    ui_main.clients_ip.clicked.connect(client_ip_config)
    ui_main.server_ip.clicked.connect(start_server_ip)
    ui_main.autoUpdatTime.timeChanged.connect(autoUpdate)
    ui_main.report.clicked.connect(start_report)
    ui_main.autoUpdateCheckBox.stateChanged.connect(resetAutoUpdateJ)
    ui_main.rpi_mode_1.setStyleSheet("background-color: transparent")
    ui_main.rpi_mode_2.setStyleSheet("background-color: transparent")
    ui_main.rpi_mode_3.setStyleSheet("background-color: transparent")
    ui_main.rpi_mode_4.setStyleSheet("background-color: transparent")
    ui_main.rpi_mode_5.setStyleSheet("background-color: transparent")
    pic = QtGui.QPixmap("offline.png")
    scene = QtWidgets.QGraphicsScene()
    scene.addPixmap(pic)
    ui_main.rpi_mode_1.setScene(scene)
    ui_main.rpi_mode_2.setScene(scene)
    ui_main.rpi_mode_3.setScene(scene)
    ui_main.rpi_mode_4.setScene(scene)
    ui_main.rpi_mode_5.setScene(scene)
    ui_main.rpi_mode_1.show()
    ui_main.rpi_mode_2.show()
    ui_main.rpi_mode_3.show()
    ui_main.rpi_mode_4.show()
    ui_main.rpi_mode_5.show()

    ui_main.comboBox.addItem("جمع کل")
    ui_main.comboBox.addItem("دستگاه یک")
    ui_main.comboBox.addItem("دستگاه دو")
    ui_main.comboBox.addItem("دستگاه سه")
    ui_main.comboBox.addItem("دستگاه چهار")
    ui_main.comboBox.addItem("دستگاه پنج")

    # Set Update Time for first Time
    ui_main.autoUpdatTime.setTime(QtCore.QTime(0,int(updateTime[0:updateTime.find(":")]),int(updateTime[updateTime.find(":")+1:])))
    
    global layout
    layout = QtWidgets.QVBoxLayout(ui_main.plot)
    ui_main.plot.setStyleSheet("background-color: transparent")


    if (firstTime):
        warn = QtWidgets.QMessageBox()
        warn.setText("تنظیمات اولیه سرور دیتابیس را وارد نمایید.")
        warn.setWindowTitle("اخطار")
        warn.exec_()
    else:
        get_server_ip()
    
    find_lastActive_table()
    update_labels()
    update_data()

    ui_main.getthread = thread_ping()
    ui_main.getthread.start()
    
    MainWindow.show()
    sys.exit(app.exec_())


create_main_folder()
get_ip_from_db()

check_default_server_ip()

prebuild()